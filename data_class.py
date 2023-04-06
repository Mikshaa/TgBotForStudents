import openpyxl


class StudentsData:

    def __init__(self):
        self.wb = openpyxl.load_workbook('data.xlsx')
        self.studList = self.wb.sheetnames
        self.idList = []
        for sheet in self.wb:
            self.idList.append([sheet.title, sheet['I2'].value])

    def getSheet(self, id):
        for pair in self.idList:
            if id == pair[1]:
                return pair[0]

    def getParName(self, id):
        return self.wb[self.getSheet(id)]['A2'].value

    def getParNum(self, id):
        return self.wb[self.getSheet(id)]['B2'].value

    def getStudNum(self, id):
        return self.wb[self.getSheet(id)]['C2'].value

    def getNextLsn(self, id):
        return str(self.wb[self.getSheet(id)]['D2'].value)[0:16]

    def getLastTheme(self, id):
        return self.wb[self.getSheet(id)]['E2'].value

    def getDz(self, id):
        return self.wb[self.getSheet(id)]['F2'].value

    def getPaid(self, id):
        return self.wb[self.getSheet(id)]['G2'].value

    def getLessons(self, id):
        return self.wb[self.getSheet(id)]['H2'].value

    def getStudName(self, id):
        return self.getSheet(id)

    def getAllStudens(self):
        return self.studList

    def getStudId(self, name):
        for i in self.idList:
            if name == i[0]:
                return i[1]


    def setParName(self, id, name):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['A2'] = name
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setParNum(self, id, num):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['B2'] = num
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setStudNum(self, id, num):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['C2'] = num
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setNextLsn(self, id, date):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['D2'] = date
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setLastTheme(self, id, theme):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['E2'] = theme
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setDz(self, id, dz):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['F2'] = dz
            self.wb.save('data.xlsx')
        else:
            print('NotFound')


    def setPaid(self, id, paid):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['G2'] = paid
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setLessons(self, id, lessons):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)]['H2'] = lessons
            self.wb.save('data.xlsx')
        else:
            print('NotFound')

    def setStudName(self, id, name):
        if self.getSheet(id) != None:
            self.wb[self.getSheet(id)].title = name
            self.studList = name
            self.wb.save('data.xlsx')
            for pair in self.idList:
                if id == pair[1]:
                    pair[0] = name
        else:
            print('NotFound')

    def newStudent(self, id, studName, parName=None, parNum=None, studNum=None, nextLsn=None, lastTheme=None, dz=None, paid=None, lessons=None):
        newSheet = self.wb.copy_worksheet(self.wb['Образец'])
        newSheet.title = studName
        self.studList.append(studName)
        self.idList.append([studName, id])
        newSheet['A2'] = parName
        newSheet['B2'] = parNum
        newSheet['C2'] = studNum
        newSheet['D2'] = nextLsn
        newSheet['E2'] = lastTheme
        newSheet['F2'] = dz
        newSheet['G2'] = paid
        newSheet['H2'] = lessons
        newSheet['I2'] = id
        self.wb.save('data.xlsx')

    def delStud(self,name):
        for i in self.wb.sheetnames:
            if name == i:
                self.wb.remove(self.wb[name])
                self.studList.remove(name)
        self.wb.save('data.xlsx')

data = StudentsData()

